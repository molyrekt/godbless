import telebot
import csv
from telebot import types
import openpyxl
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
from io import BytesIO

bot = telebot.TeleBot('6709849437:AAGM6WS74oD7CvdYK5l-biN_40hBExi_WX0')

def iphone_models_pricing(callback):
    bot.delete_message(callback.message.chat.id, callback.message.message_id)
    keyboard = types.InlineKeyboardMarkup()
    with open('pricing_bot.csv',mode='r', encoding='utf-8') as csvfile:
        reader = list(csv.reader(csvfile))
        models = reader[2:]
        for row in models:
            callback_data, model = row[1], row[2]
            button = types.InlineKeyboardButton(text=f'iPhone {model} 🔥', callback_data=callback_data)
            keyboard.add(button)
    file_path = './ppg_repair.png'
    with open(file_path, 'rb') as file:
        bot.send_photo(callback.message.chat.id, file, reply_markup=keyboard)



def iphone_models_menu(callback, page):
    bot.delete_message(callback.message.chat.id, callback.message.message_id)
    keyboard = types.InlineKeyboardMarkup()
    models_per_page = 15
    start_index = (page - 1) * models_per_page

    with open('repair_bot.csv', mode='r', encoding='utf-8') as csvfile:
        reader = list(csv.reader(csvfile))
        models = reader[1:]
        for row in models[start_index:start_index + models_per_page]:
            callback_data, model = row[0], row[1]
            button = types.InlineKeyboardButton(text=f'iPhone {model} 🔥', callback_data=callback_data)
            keyboard.add(button)

        if page == 1 and len(models) > models_per_page:
            forward_button = types.InlineKeyboardButton(text="Вперед ➡️", callback_data="iphone_models_next")
            keyboard.add(forward_button)
        elif page == 2:
            back_button = types.InlineKeyboardButton(text="⬅️ Назад", callback_data="iphone_models_back")
            keyboard.add(back_button)

    file_path = './ppg_repair.png'
    with open(file_path, 'rb') as file:
        bot.send_photo(callback.message.chat.id, file, reply_markup=keyboard)





def display_services_menu(callback, model_name):
    keyboard = types.InlineKeyboardMarkup()
    with open('repair_bot.csv', mode='r', encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            if row['МОДЕЛЬ'] == model_name:
                for service, price in list(row.items())[2:]:
                    if price != 'не обслуживается':
                        button_text = f"{service}: {price} ₸"
                        button = types.InlineKeyboardButton(text=button_text, callback_data=f"service_{service}")
                        keyboard.add(button)
                break
    back_button = types.InlineKeyboardButton(text="⬅️ Назад️", callback_data="back_to_models")
    keyboard.add(back_button)

    file_path = './ppg_repair.png'
    with open(file_path, 'rb') as file:
        bot.send_photo(callback.message.chat.id, file, reply_markup=keyboard)






def main_menu_tradein(callback):
    file_path = './ppg_tradein.png'
    with open(file_path, 'rb') as image_file:
        markup = types.InlineKeyboardMarkup()

        wb = openpyxl.load_workbook('tradein.xlsx', data_only=True)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            _, button_text, callback_data = row[:3]
            price = row[10]
            button = types.InlineKeyboardButton(f'{button_text}{price}', callback_data=callback_data)
            markup.add(button)

        bot.send_photo(callback.message.chat.id, image_file, reply_markup=markup)





def first_page_iphones_pricing(callback):
    file_path1 = './ppg_pricing.png'
    with open(file_path1, 'rb') as file1:
        keyboard = types.InlineKeyboardMarkup()

        iphone_btn1 = types.InlineKeyboardButton(text='14 Pro Max 🔥', callback_data='pricing_14pm')
        iphone_btn2 = types.InlineKeyboardButton(text='14 Pro 🔥', callback_data='pricing_14pro')
        iphone_btn3 = types.InlineKeyboardButton(text='14 Plus 🔥', callback_data='pricing_14plus')
        iphone_btn4 = types.InlineKeyboardButton(text='14 🔥', callback_data='pricing_14')
        iphone_btn5 = types.InlineKeyboardButton(text='13PM 🔥', callback_data='pricing_13pm')
        iphone_btn6 = types.InlineKeyboardButton(text='13 Pro 🔥', callback_data='pricing_13pro')
        iphone_btn7 = types.InlineKeyboardButton(text='13 mini 🔥', callback_data='pricing_13mini')
        iphone_btn8 = types.InlineKeyboardButton(text='13 🔥', callback_data='pricing_13')
        iphone_btn9 = types.InlineKeyboardButton(text='12 Pro Max 🔥', callback_data='pricing_12pm')
        iphone_btn10 = types.InlineKeyboardButton(text='12 Pro 🔥', callback_data='pricing_12pro')
        iphone_btn11 = types.InlineKeyboardButton(text='12 mini 🔥', callback_data='pricing_12mini')
        iphone_btn12 = types.InlineKeyboardButton(text='12 🔥', callback_data='pricing_12')
        iphone_btn13 = types.InlineKeyboardButton(text='◀️ Назад', callback_data='pricing_previous_page1')
        iphone_btn14 = types.InlineKeyboardButton(text='Вперед ➡️', callback_data='pricing_next_page1')

        keyboard.add(iphone_btn1, iphone_btn2)
        keyboard.add(iphone_btn3, iphone_btn4)
        keyboard.add(iphone_btn5, iphone_btn6)
        keyboard.add(iphone_btn7, iphone_btn8)
        keyboard.add(iphone_btn9, iphone_btn10)
        keyboard.add(iphone_btn11, iphone_btn12)
        keyboard.add(iphone_btn13, iphone_btn14)

        bot.send_photo(callback.message.chat.id, file1)
        bot.send_message(callback.message.chat.id, 'Список всех устройств, которые Вы можете продать в PaPaGadget 🔥',
                         reply_markup=keyboard)


#### ОЦЕНКА 14 про макс начало



def pricing_14pm_0step(callback):
    bot.delete_message(callback.message.chat.id, callback.message.message_id)
    bot.delete_message(callback.message.chat.id, callback.message.message_id - 1)
    file_path = './ppg_tradein.png'
    with open(file_path, 'rb') as file:
        keyboard = types.InlineKeyboardMarkup()

        korpus_14pm_btn1 = types.InlineKeyboardButton(text='В идеале 🔥', callback_data='14pm_korpus_ideal')
        korpus_14pm_btn2 = types.InlineKeyboardButton(text='Немного потертый ✅', callback_data='14pm_korpus_medium')
        korpus_14pm_btn3 = types.InlineKeyboardButton(text='В царапинах / поношенный 💀',
                                                      callback_data='14pm_korpus_bad')

        keyboard.add(korpus_14pm_btn1)
        keyboard.add(korpus_14pm_btn2)
        keyboard.add(korpus_14pm_btn3)

        bot.send_photo(callback.message.chat.id, file)
        bot.send_message(callback.message.chat.id, 'В каком состоянии Ваш корпус? 🔥', reply_markup=keyboard)



########### ОЦЕНКА КОРПУСА 14 ПМ ПЕРЕХОД НА ОЦЕНКУ ДИСПЛЕЯ #######
def pricing_14pm_1step(callback):
    bot.delete_message(callback.message.chat.id, callback.message.message_id)
    bot.delete_message(callback.message.chat.id, callback.message.message_id - 1)
    file_path = './ppg_pricing.png'
    with open(file_path, 'rb') as file:
        keyboard = types.InlineKeyboardMarkup()
        screen_14pm_btn1 = types.InlineKeyboardButton(text='В идеале 🔥', callback_data='14pm_display_ideal')
        screen_14pm_btn2 = types.InlineKeyboardButton(text='Царапины на экране ✅', callback_data='14pm_display_scratch')
        screen_14pm_btn3 = types.InlineKeyboardButton(text='Разбито стекло дисплея 💀',
                                                      callback_data='14pm_display_badglass')
        screen_14pm_btn4 = types.InlineKeyboardButton(text='Полосы / пятна на экране 💀',
                                                      callback_data='14pm_display_badlines')
        screen_14pm_btn5 = types.InlineKeyboardButton(text='Экран не работает 💀', callback_data='14pm_display_fucked')

        keyboard.add(screen_14pm_btn1)
        keyboard.add(screen_14pm_btn2)
        keyboard.add(screen_14pm_btn3)
        keyboard.add(screen_14pm_btn4)
        keyboard.add(screen_14pm_btn5)

        bot.send_photo(callback.message.chat.id, file)
        bot.send_message(callback.message.chat.id, 'В каком состоянии Ваш экран? 🔥', reply_markup=keyboard)



####### ОЦЕНКА ДИСПЛЕЯ 14 ПМ ПЕРЕХОД НА ОЦЕНКУ КАМЕР #######

def pricing_14pm_2step(callback):
    bot.delete_message(callback.message.chat.id, callback.message.message_id)
    bot.delete_message(callback.message.chat.id, callback.message.message_id - 1)
    file_path='./ppg_pricing.png'
    with open(file_path, 'rb') as file:
        keyboard=types.InlineKeyboardMarkup()

        camera_14pm_btn1 = types.InlineKeyboardButton(text='В идеале 🔥', callback_data='14pm_camera_ideal')
        camera_14pm_btn2 = types.InlineKeyboardButton(text='Разбито стекло камеры ✅', callback_data='14pm_camera_badglass')
        camera_14pm_btn3 = types.InlineKeyboardButton(text='Не работает камера 💀', callback_data='14pm_camera_fucked')

        keyboard.add(camera_14pm_btn1)
        keyboard.add(camera_14pm_btn2)
        keyboard.add(camera_14pm_btn3)

        bot.send_photo(callback.message.chat.id,file)
        bot.send_message(callback.message.chat.id, 'В каком состоянии Ваша камера? 🔥', reply_markup=keyboard)



######## Оценка других проблем 14 ПМ



def pricing_14pm_3step(callback):
    bot.delete_message(callback.message.chat.id, callback.message.message_id)
    bot.delete_message(callback.message.chat.id, callback.message.message_id - 1)
    file_path='./ppg_pricing.png'
    with open(file_path, 'rb') as file:
        keyboard=types.InlineKeyboardMarkup()

        other_14pm_btn1 = types.InlineKeyboardButton(text='Других проблем нет 🔥', callback_data='14pm_other_ideal')
        other_14pm_btn2 = types.InlineKeyboardButton(text='Не работает Face ID ✅', callback_data='14pm_other_faceid')
        other_14pm_btn3 = types.InlineKeyboardButton(text='Проблемы со звуком ✅', callback_data='14pm_other_sound')
        other_14pm_btn4 = types.InlineKeyboardButton(text='Нерабочие кнопки ✅', callback_data='14pm_other_buttons')

        keyboard.add(other_14pm_btn1)
        keyboard.add(other_14pm_btn2)
        keyboard.add(other_14pm_btn3)
        keyboard.add(other_14pm_btn4)

        bot.send_photo(callback.message.chat.id,file)
        bot.send_message(callback.message.chat.id, 'Есть ли другие неисправности в телефоне? 🔥', reply_markup=keyboard)



###### ФИНАЛ ОЦЕНКИ 14ПМ

def pricing_14pm_4step(callback):
    bot.delete_message(callback.message.chat.id, callback.message.message_id)
    bot.delete_message(callback.message.chat.id, callback.message.message_id - 1)
    file_path = './ppg_pricing.png'
    with open(file_path, 'rb') as file:
        keyboard = types.InlineKeyboardMarkup()

        final_14pm_btn1 = types.InlineKeyboardButton(text='Ваш iPhone 14 Pro Max оценён :', callback_data='final_14pm_description')
        final_14pm_btn2 = types.InlineKeyboardButton(text=f'{price14pm} ✅', callback_data='final_14pm_price')
        final_14pm_btn3 = types.InlineKeyboardButton(text='В главное меню 🔙', callback_data='back_main_menu_14pm_pricing')

        keyboard.add(final_14pm_btn1)
        keyboard.add(final_14pm_btn2)
        keyboard.add(final_14pm_btn3)

        bot.send_photo(callback.message.chat.id, file, reply_markup=keyboard)




        ####### Оценка 14 про начало

def pricing_14pro_0step(callback):
    bot.delete_message(callback.message.chat.id, callback.message.message_id)
    bot.delete_message(callback.message.chat.id, callback.message.message_id - 1)
    file_path = './ppg_tradein.png'
    with open(file_path, 'rb') as file:
        keyboard = types.InlineKeyboardMarkup()

        korpus_14pm_btn1 = types.InlineKeyboardButton(text='В идеале 🔥', callback_data='14pro_korpus_ideal')
        korpus_14pm_btn2 = types.InlineKeyboardButton(text='Немного потертый ✅', callback_data='14pro_korpus_medium')
        korpus_14pm_btn3 = types.InlineKeyboardButton(text='В царапинах / поношенный 💀',
                                                      callback_data='14pro_korpus_bad')

        keyboard.add(korpus_14pm_btn1)
        keyboard.add(korpus_14pm_btn2)
        keyboard.add(korpus_14pm_btn3)

        bot.send_photo(callback.message.chat.id, file)
        bot.send_message(callback.message.chat.id, 'В каком состоянии Ваш корпус? 🔥', reply_markup=keyboard)



#####ОЦЕНКА 14 про дисплей

def pricing_14pro_1step(callback):
    bot.delete_message(callback.message.chat.id, callback.message.message_id)
    bot.delete_message(callback.message.chat.id, callback.message.message_id - 1)
    file_path = './ppg_pricing.png'
    with open(file_path, 'rb') as file:
        keyboard = types.InlineKeyboardMarkup()
        screen_14pro_btn1 = types.InlineKeyboardButton(text='В идеале 🔥', callback_data='14pro_display_ideal')
        screen_14pro_btn2 = types.InlineKeyboardButton(text='Царапины на экране ✅', callback_data='14pro_display_scratch')
        screen_14pro_btn3 = types.InlineKeyboardButton(text='Разбито стекло дисплея 💀',
                                                      callback_data='14pro_display_badglass')
        screen_14pro_btn4 = types.InlineKeyboardButton(text='Полосы / пятна на экране 💀',
                                                      callback_data='14pro_display_badlines')
        screen_14pro_btn5 = types.InlineKeyboardButton(text='Экран не работает 💀', callback_data='14pro_display_fucked')

        keyboard.add(screen_14pro_btn1)
        keyboard.add(screen_14pro_btn2)
        keyboard.add(screen_14pro_btn3)
        keyboard.add(screen_14pro_btn4)
        keyboard.add(screen_14pro_btn5)

        bot.send_photo(callback.message.chat.id, file)
        bot.send_message(callback.message.chat.id, 'В каком состоянии Ваш экран? 🔥', reply_markup=keyboard)



#### ОЦЕНКА 14 ПРО камеры

def pricing_14pro_2step(callback):
    bot.delete_message(callback.message.chat.id, callback.message.message_id)
    bot.delete_message(callback.message.chat.id, callback.message.message_id - 1)
    file_path='./ppg_pricing.png'
    with open(file_path, 'rb') as file:
        keyboard=types.InlineKeyboardMarkup()

        camera_14pro_btn1 = types.InlineKeyboardButton(text='В идеале 🔥', callback_data='14pro_camera_ideal')
        camera_14pro_btn2 = types.InlineKeyboardButton(text='Разбито стекло камеры ✅', callback_data='14pro_camera_badglass')
        camera_14pro_btn3 = types.InlineKeyboardButton(text='Не работает камера 💀', callback_data='14pro_camera_fucked')

        keyboard.add(camera_14pro_btn1)
        keyboard.add(camera_14pro_btn2)
        keyboard.add(camera_14pro_btn3)

        bot.send_photo(callback.message.chat.id,file)
        bot.send_message(callback.message.chat.id, 'В каком состоянии Ваша камера? 🔥', reply_markup=keyboard)


###### ОЦЕНКА 14 ПРО ДРУГИЕ ПРОБЛЕМЫ

def pricing_14pro_3step(callback):
    bot.delete_message(callback.message.chat.id, callback.message.message_id)
    bot.delete_message(callback.message.chat.id, callback.message.message_id - 1)
    file_path='./ppg_pricing.png'
    with open(file_path, 'rb') as file:
        keyboard=types.InlineKeyboardMarkup()

        other_14pro_btn1 = types.InlineKeyboardButton(text='Других проблем нет 🔥', callback_data='14pro_other_ideal')
        other_14pro_btn2 = types.InlineKeyboardButton(text='Не работает Face ID ✅', callback_data='14pro_other_faceid')
        other_14pro_btn3 = types.InlineKeyboardButton(text='Проблемы со звуком ✅', callback_data='14pro_other_sound')
        other_14pro_btn4 = types.InlineKeyboardButton(text='Нерабочие кнопки ✅', callback_data='14pro_other_buttons')

        keyboard.add(other_14pro_btn1)
        keyboard.add(other_14pro_btn2)
        keyboard.add(other_14pro_btn3)
        keyboard.add(other_14pro_btn4)

        bot.send_photo(callback.message.chat.id,file)
        bot.send_message(callback.message.chat.id, 'Есть ли другие неисправности в телефоне? 🔥', reply_markup=keyboard)


#### ОЦЕНКА 14 ПРО ФИНАЛ

def pricing_14pro_4step(callback):
    bot.delete_message(callback.message.chat.id, callback.message.message_id)
    bot.delete_message(callback.message.chat.id, callback.message.message_id - 1)
    file_path = './ppg_pricing.png'
    with open(file_path, 'rb') as file:
        keyboard = types.InlineKeyboardMarkup()

        final_14pro_btn1 = types.InlineKeyboardButton(text='Ваш iPhone 14 Pro оценён :', callback_data='final_14pro_description')
        final_14pro_btn2 = types.InlineKeyboardButton(text=f'{price14pm} ✅', callback_data='final_14pro_price')
        final_14pro_btn3 = types.InlineKeyboardButton(text='В главное меню 🔙', callback_data='back_main_menu_14pro_pricing')

        keyboard.add(final_14pro_btn1)
        keyboard.add(final_14pro_btn2)
        keyboard.add(final_14pro_btn3)

        bot.send_photo(callback.message.chat.id, file, reply_markup=keyboard)

        ####### REPLY КЛАВИАТУРА


@bot.message_handler(commands=['start'])
def start(message):

    #Основные кнопки

    markup = types.ReplyKeyboardMarkup()
    btn1 = types.KeyboardButton('Trade In ♻️')
    btn2 = types.KeyboardButton('Ремонт 🛠')
    btn3 = types.KeyboardButton('Оценка устройства 💼')
    btn4 = types.KeyboardButton('Аксессуары ☎️')
    btn5 = types.KeyboardButton('Original 🍏')
    btn6 = types.KeyboardButton('Связаться с нами 💬')
    markup.row(btn1, btn2)
    markup.row(btn4,btn5)
    markup.row(btn3, btn6)




    #Приветствие

    bot.send_message(message.chat.id,'Доброго времени суток уважаемый Клиент! Этот бот - Ваш личный помощник в сервисе PaPaGadget 🔥', reply_markup=markup)


###### СВЯЗАТЬСЯ С НАМИ #####
@bot.message_handler(func=lambda message:message.text == 'Связаться с нами 💬')
def support(message):
    if message.text=='Связаться с нами 💬':
        file_path = './ppg_contactus.png'
        with open(file_path, 'rb') as file:
            markup = types.InlineKeyboardMarkup()
            support_btn = types.InlineKeyboardButton('Напишите нам 💬', url='https://t.me/molyrekt')
            markup.add(support_btn)
            bot.send_photo(message.chat.id,file, reply_markup=markup)




#Trade In Menu


@bot.message_handler(func=lambda message: message.text == 'Trade In ♻️')
def tradein(message):
    if message.text == 'Trade In ♻️':
        file_path = './ppg_tradein.png'
        with open(file_path, 'rb') as image_file:
            markup = types.InlineKeyboardMarkup()

            wb = openpyxl.load_workbook('tradein.xlsx', data_only=True)
            ws = wb.active

            for row in ws.iter_rows(min_row=1, values_only=True):
                _, button_text, callback_data = row[:3]
                price = row[9]
                button = types.InlineKeyboardButton(f'{button_text}{price}', callback_data=callback_data)
                markup.add(button)

            bot.send_photo(message.chat.id, image_file, reply_markup=markup)



price14pm = 500000
price14pro = 480000




############Tradein MENU


wb = openpyxl.load_workbook('tradein.xlsx', data_only=True)
ws = wb.active

devices_info = []

image_loader = SheetImageLoader(ws)

for index, row in enumerate(ws.iter_rows(values_only=True), start=1):
    device_info = {
        "name": row[1],
        "callback_data": row[2],
        "buttons": [
            {"text": row[3], "callback_data": "x"},
            {"text": row[4], "callback_data": "x"},
            {"text": row[5], "callback_data": "x"},
            {"text": row[6], "callback_data": "x"}
        ],
        "price": row[9],
        "back_button": {"text": "Назад ◀️", "callback_data": "back_tradein"},
        "images": []
    }

    # Attempt to load images from cells H and I; handle ValueError if no image is found
    for cell in ['H', 'I']:
        try:
            img = image_loader.get(f'{cell}{index}')
            if img:
                device_info["images"].append(img)
        except ValueError:
            # No image in this cell, continue to the next cell
            continue

    devices_info.append(device_info)


@bot.callback_query_handler(func=lambda callback: callback.data)
def callback_tradein(callback):
    for device in devices_info:
        if callback.data == device["callback_data"]:
            bot.delete_message(callback.message.chat.id, callback.message.message_id)
            keyboard = types.InlineKeyboardMarkup()

            # Добавление кнопок для устройства
            for btn_info in device["buttons"]:
                btn = types.InlineKeyboardButton(text=btn_info["text"], callback_data=btn_info["callback_data"])
                keyboard.add(btn)

            # Добавление кнопки 'Назад'
            back_btn_info = device["back_button"]
            back_btn = types.InlineKeyboardButton(text=back_btn_info["text"], callback_data='back_tradein')
            keyboard.add(back_btn)

            #отправка изображений
            media = [types.InputMediaPhoto(media=img) for img in device["images"] if img is not None]

            bot.send_media_group(callback.message.chat.id, media=media)
            bot.send_message(callback.message.chat.id, device["name"], reply_markup=keyboard)


    global price14pm, price14pro




############Возвращение в трейд ин меню


    if callback.data=='back_tradein':
        bot.delete_message(callback.message.chat.id, callback.message.message_id)
        bot.delete_message(callback.message.chat.id, callback.message.message_id - 1)
        bot.delete_message(callback.message.chat.id, callback.message.message_id - 2)

        main_menu_tradein(callback)


        # СПИСОК IPHONE УСТРОЙСТВ





    if callback.data=='repair_main_menu':
        file_path1='./ppg_repair.png'
        bot.delete_message(callback.message.chat.id, callback.message.message_id)
        bot.delete_message(callback.message.chat.id, callback.message.message_id - 1)
        with open(file_path1, 'rb') as file:
            markup = types.InlineKeyboardMarkup()
            repair_btn2 = types.InlineKeyboardButton('iPhone 🔥', callback_data='iphone')
            repair_btn3 = types.InlineKeyboardButton('Samsung 🔥', callback_data='samsung')
            repair_btn4 = types.InlineKeyboardButton('Xiaomi 🔥', callback_data='xiaomi')
            repair_btn5 = types.InlineKeyboardButton('Poco 🔥', callback_data='poco')
            markup.row(repair_btn2, repair_btn3)
            markup.row(repair_btn4, repair_btn5)
            bot.send_photo(callback.message.chat.id, file)
            bot.send_message(callback.message.chat.id, '📱 Выберите Ваше устройство 📱',
                             reply_markup=markup)


    if callback.data.startswith('repair_'):
        model_name = callback.data.split('_')[1]
        display_services_menu(callback, model_name)

    elif callback.data == 'iphone_models_next':
        iphone_models_menu(callback, 2)

    elif callback.data == 'iphone_models_back':
        iphone_models_menu(callback, 1)

    elif callback.data == 'iphone':
        iphone_models_menu(callback, 1)

    if callback.data == 'back_to_models':
        iphone_models_menu(callback,1)

    if callback.data=='pricing_14pm':
        pricing_14pm_0step(callback)


##### ОЦЕНКА КОРПУСА 14 ПМ ИДЕАЛ


    if callback.data == '14pm_korpus_ideal':
        price14pm-=0
        pricing_14pm_1step(callback)


###### ОЦЕНКА КОРПУСА 14 ПМ МЕДИУМ


    elif callback.data == '14pm_korpus_medium':
        price14pm-=40000
        pricing_14pm_1step(callback)


###### ОЦЕНКА КОРПУСА 14 ПМ УГАШЕН


    elif callback.data == '14pm_korpus_bad':
        price14pm-=65000
        pricing_14pm_1step(callback)




###### ОЦЕНКА ДИСПЛЕЯ 14 ПМ ИДЕАЛ

    elif callback.data == '14pm_display_ideal':
        price14pm-=0
        pricing_14pm_2step(callback)



##### ОЦЕНКА ДИСПЛЕЯ 14 ПМ МЕДИУМ

    elif callback.data == '14pm_display_scratch':
        price14pm-=10000
        pricing_14pm_2step(callback)


###### ОЦЕНКА ДИСПЛЕЯ 14 ПМ РАЗБИТО СТЕКЛО

    elif callback.data == '14pm_display_badglass':
        price14pm-=30000
        pricing_14pm_2step(callback)


####### ОЦЕНКА ДИСПЛЕЯ 14 ПМ ПЯТНА ПОЛОСЫ

    elif callback.data == '14pm_display_badlines':
        price14pm-=200000
        pricing_14pm_2step(callback)


###### ОЦЕНКА ДИСПЛЕЯ 14 ПМ НЕ РАБОТАЕТ ЭКРАН

    elif callback.data == '14pm_display_fucked':
        price14pm-=230000
        pricing_14pm_2step(callback)

###### ОЦЕНКА КАМЕРЫ 14 ПМ ИДЕАЛ

    elif callback.data == '14pm_camera_ideal':
        price14pm-=0
        pricing_14pm_3step(callback)

###### ОЦЕНКА КАМЕРЫ 14 ПМ РАЗБИТО СТЕКЛО КАМЕРЫ

    elif callback.data == '14pm_camera_badglass':
        price14pm-=10000
        pricing_14pm_3step(callback)

####### ОЦЕНКА КАМЕРЫ 14ПМ НЕ РАБОТАЕТ КАМЕРА

    elif callback.data == '14pm_camera_fucked':
        price14pm-=25000
        pricing_14pm_3step(callback)


####### Оценка других проблем 14 пм Нет проблем

    elif callback.data == '14pm_other_ideal':
        price14pm-=0
        pricing_14pm_4step(callback)

###### Оценка других проблем 14 пм face id

    elif callback.data == '14pm_other_faceid':
        price14pm-=30000
        pricing_14pm_4step(callback)

        ######Оценка других проблем 14пм проблемы со звуком

    elif callback.data == '14pm_other_sound':
        price14pm-=15000
        pricing_14pm_4step(callback)


        ###### Оценка других проблем 14пм проблемы с кнопками

    elif callback.data == '14pm_other_buttons':
        price14pm-=15000
        pricing_14pm_4step(callback)

####### Возврат в главное меню из 14пм

    elif callback.data == 'back_main_menu_14pm_pricing':
        bot.delete_message(callback.message.chat.id, callback.message.message_id)
        first_page_iphones_pricing(callback)
        price14pm=500000





    ######## Оценка 14 про макс


    if callback.data=='pricing_14pro':
        pricing_14pro_0step(callback)


##### ОЦЕНКА КОРПУСА 14 Про ИДЕАЛ


    if callback.data == '14pro_korpus_ideal':
        price14pro-=0
        pricing_14pro_1step(callback)


###### ОЦЕНКА КОРПУСА 14 Про МЕДИУМ


    elif callback.data == '14pro_korpus_medium':
        price14pro-=42000
        pricing_14pro_1step(callback)


###### ОЦЕНКА КОРПУСА 14 Про УГАШЕН


    elif callback.data == '14pro_korpus_bad':
        price14pro-=60000
        pricing_14pro_1step(callback)




###### ОЦЕНКА ДИСПЛЕЯ 14 Про ИДЕАЛ

    elif callback.data == '14pro_display_ideal':
        price14pro-=0
        pricing_14pro_2step(callback)



##### ОЦЕНКА ДИСПЛЕЯ 14 Про МЕДИУМ

    elif callback.data == '14pro_display_scratch':
        price14pro-=12000
        pricing_14pro_2step(callback)


###### ОЦЕНКА ДИСПЛЕЯ 14 Про РАЗБИТО СТЕКЛО

    elif callback.data == '14pro_display_badglass':
        price14pro-=35000
        pricing_14pro_2step(callback)


####### ОЦЕНКА ДИСПЛЕЯ 14 Про ПЯТНА ПОЛОСЫ

    elif callback.data == '14pro_display_badlines':
        price14pro-=210000
        pricing_14pro_2step(callback)


###### ОЦЕНКА ДИСПЛЕЯ 14 Про НЕ РАБОТАЕТ ЭКРАН

    elif callback.data == '14pro_display_fucked':
        price14pro-=220000
        pricing_14pro_2step(callback)

###### ОЦЕНКА КАМЕРЫ 14 Про ИДЕАЛ

    elif callback.data == '14pro_camera_ideal':
        price14pro-=0
        pricing_14pro_3step(callback)

###### ОЦЕНКА КАМЕРЫ 14 Про РАЗБИТО СТЕКЛО КАМЕРЫ

    elif callback.data == '14pro_camera_badglass':
        price14pro-=12000
        pricing_14pro_3step(callback)

####### ОЦЕНКА КАМЕРЫ 14Про НЕ РАБОТАЕТ КАМЕРА

    elif callback.data == '14pro_camera_fucked':
        price14pro-=20000
        pricing_14pro_3step(callback)


####### Оценка других проблем 14 про Нет проблем

    elif callback.data == '14pro_other_ideal':
        price14pro-=0
        pricing_14pro_4step(callback)

###### Оценка других проблем 14 про face id

    elif callback.data == '14pro_other_faceid':
        price14pro-=25000
        pricing_14pro_4step(callback)

        ######Оценка других проблем 14про проблемы со звуком

    elif callback.data == '14pro_other_sound':
        price14pro-=10000
        pricing_14pro_4step(callback)


        ###### Оценка других проблем 14про проблемы с кнопками

    elif callback.data == '14pro_other_buttons':
        price14pro-=10000
        pricing_14pro_4step(callback)

####### Возврат в главное меню из 14про

    elif callback.data == 'back_main_menu_14pro_pricing':
        bot.delete_message(callback.message.chat.id, callback.message.message_id)
        first_page_iphones_pricing(callback)
        price14pro=480000


########### Ремонт - основной раздел

@bot.message_handler(func=lambda message: message.text == 'Ремонт 🛠')
def repair(message):
    if message.text == "Ремонт 🛠":
        file_path = './ppg_repair.png'
        with open(file_path, 'rb') as file:
            markup = types.InlineKeyboardMarkup()
            repair_btn1 = types.InlineKeyboardButton('Выберите Ваше устройство 📱',callback_data='choose')
            repair_btn2 = types.InlineKeyboardButton('iPhone 🔥', callback_data='iphone')
            repair_btn3 = types.InlineKeyboardButton('Samsung 🔥', callback_data='samsung')
            repair_btn4 = types.InlineKeyboardButton('Xiaomi 🔥', callback_data='xiaomi')
            repair_btn5 = types.InlineKeyboardButton('Poco 🔥', callback_data='poco')
            markup.row(repair_btn1)
            markup.row(repair_btn2,repair_btn3)
            markup.row(repair_btn4,repair_btn5)
            bot.send_photo(message.chat.id,file,reply_markup=markup)
        file.close()


@bot.message_handler(func=lambda message: message.text == 'Оценка устройства 💼')
def iphone_pricing(message):
    if message.text == "Оценка устройства 💼":
        bot.delete_message(message.chat.id, message.message_id)
        keyboard = types.InlineKeyboardMarkup()
        with open('pricing_bot.csv', mode='r', encoding='utf-8') as csvfile:
            reader = list(csv.reader(csvfile))
            models = reader[2:]
            for row in models:
                callback_data, model = row[1], row[2]
                button = types.InlineKeyboardButton(text=f'iPhone {model} 🔥', callback_data=callback_data)
                keyboard.add(button)
        file_path = './ppg_repair.png'
        with open(file_path, 'rb') as file:
            bot.send_photo(message.chat.id, file, reply_markup=keyboard)


bot.polling(none_stop=True)